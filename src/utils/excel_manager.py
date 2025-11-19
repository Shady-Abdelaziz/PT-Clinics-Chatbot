"""
Excel Database Manager
Handles all appointment operations: viewing, booking, and canceling
"""
import pandas as pd
from datetime import datetime, timedelta
from pathlib import Path
from typing import List, Dict, Optional, Tuple
import openpyxl
from openpyxl.styles import Font, PatternFill


class ExcelDBManager:
    """Manages the Excel database for appointments"""
    
    def __init__(self, excel_path: str):
        """Initialize Excel DB Manager"""
        self.excel_path = Path(excel_path)
        if not self.excel_path.exists():
            raise FileNotFoundError(f"Excel database not found at: {excel_path}")
        
        # Load all sheets
        self.excel_file = pd.ExcelFile(self.excel_path)
        self.sheet_names = self.excel_file.sheet_names
        
        # Doctor names (all sheets except 'Patients')
        self.doctor_sheets = [name for name in self.sheet_names if name != 'Patients']
    
    def get_all_doctors(self) -> List[str]:
        """Get list of all doctors"""
        return self.doctor_sheets.copy()
    
    def get_doctor_info(self, doctor_name: str) -> Optional[Dict]:
        """Get detailed information about a specific doctor"""
        # This would come from the vector DB, but we can provide basic info
        return {
            "name": doctor_name,
            "available": doctor_name in self.doctor_sheets
        }
    
    def get_available_slots(
        self, 
        doctor_name: str, 
        date: Optional[str] = None,
        limit: int = 10
    ) -> List[Dict]:
        """
        Get available appointment slots for a doctor
        
        Args:
            doctor_name: Name of the doctor
            date: Specific date (YYYY-MM-DD) or None for all upcoming
            limit: Maximum number of slots to return
            
        Returns:
            List of available slots with date, time, and doctor info
        """
        if doctor_name not in self.doctor_sheets:
            return []
        
        # Read doctor's schedule
        df = pd.read_excel(self.excel_path, sheet_name=doctor_name)
        
        # Filter for available slots
        available_df = df[df['Status'] == 'Available'].copy()
        
        # Convert Date column to datetime
        available_df['Date'] = pd.to_datetime(available_df['Date'])
        
        # Filter by date if specified
        if date:
            target_date = pd.to_datetime(date)
            available_df = available_df[available_df['Date'] == target_date]
        else:
            # Only show future dates
            today = pd.Timestamp.now().normalize()
            available_df = available_df[available_df['Date'] >= today]
        
        # Sort by date and time
        available_df = available_df.sort_values(['Date', 'Time'])
        
        # Limit results
        available_df = available_df.head(limit)
        
        # Format results
        results = []
        for _, row in available_df.iterrows():
            results.append({
                'doctor': doctor_name,
                'date': row['Date'].strftime('%Y-%m-%d'),
                'time': row['Time'],
                'status': row['Status']
            })
        
        return results
    
    def book_appointment(
        self,
        doctor_name: str,
        date: str,
        time: str,
        patient_name: str,
        phone: str
    ) -> Tuple[bool, str]:
        """
        Book an appointment
        
        Args:
            doctor_name: Name of the doctor
            date: Appointment date (YYYY-MM-DD)
            time: Appointment time (HH:MM AM/PM)
            patient_name: Patient's full name
            phone: Patient's phone number
            
        Returns:
            Tuple of (success: bool, message: str)
        """
        if doctor_name not in self.doctor_sheets:
            return False, f"Doctor '{doctor_name}' not found in the system."
        
        try:
            # Load the workbook
            wb = openpyxl.load_workbook(self.excel_path)
            ws = wb[doctor_name]
            
            # Find the matching row
            target_date = datetime.strptime(date, '%Y-%m-%d').strftime('%Y-%m-%d')
            found = False
            row_index = None
            
            for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
                cell_date = row[0].value
                cell_time = row[1].value
                cell_status = row[4].value
                
                # Convert cell date to string for comparison
                if isinstance(cell_date, datetime):
                    cell_date_str = cell_date.strftime('%Y-%m-%d')
                else:
                    cell_date_str = str(cell_date)
                
                if cell_date_str == target_date and str(cell_time) == time and cell_status == 'Available':
                    found = True
                    row_index = idx
                    break
            
            if not found:
                return False, f"No available slot found for {doctor_name} on {date} at {time}"
            
            # Update the row
            ws.cell(row=row_index, column=3, value=patient_name)  # Patient_Name
            ws.cell(row=row_index, column=4, value=phone)  # Phone
            ws.cell(row=row_index, column=5, value='Reserved')  # Status
            
            # Apply formatting
            ws.cell(row=row_index, column=5).fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
            
            # Save the workbook
            wb.save(self.excel_path)
            wb.close()
            
            return True, f"✅ Appointment booked successfully!\n\nDoctor: {doctor_name}\nDate: {date}\nTime: {time}\nPatient: {patient_name}\nPhone: {phone}"
            
        except Exception as e:
            return False, f"Error booking appointment: {str(e)}"
    
    def cancel_appointment(
        self,
        doctor_name: str,
        patient_name: str,
        date: Optional[str] = None,
        time: Optional[str] = None
    ) -> Tuple[bool, str]:
        """
        Cancel an appointment
        
        Args:
            doctor_name: Name of the doctor
            patient_name: Patient's name
            date: Appointment date (optional)
            time: Appointment time (optional)
            
        Returns:
            Tuple of (success: bool, message: str)
        """
        if doctor_name not in self.doctor_sheets:
            return False, f"Doctor '{doctor_name}' not found in the system."
        
        try:
            # Load the workbook
            wb = openpyxl.load_workbook(self.excel_path)
            ws = wb[doctor_name]
            
            # Find the matching row
            found = False
            cancelled_appointments = []
            
            for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
                cell_date = row[0].value
                cell_time = row[1].value
                cell_patient = row[2].value
                cell_status = row[4].value
                
                # Convert cell date to string for comparison
                if isinstance(cell_date, datetime):
                    cell_date_str = cell_date.strftime('%Y-%m-%d')
                else:
                    cell_date_str = str(cell_date)
                
                # Check if this is the appointment to cancel
                matches_patient = cell_patient == patient_name
                matches_status = cell_status == 'Reserved'
                matches_date = (date is None) or (cell_date_str == date)
                matches_time = (time is None) or (str(cell_time) == time)
                
                if matches_patient and matches_status and matches_date and matches_time:
                    # Cancel the appointment
                    ws.cell(row=idx, column=3, value='-')  # Clear Patient_Name
                    ws.cell(row=idx, column=4, value='-')  # Clear Phone
                    ws.cell(row=idx, column=5, value='Available')  # Status
                    
                    # Remove formatting
                    ws.cell(row=idx, column=5).fill = PatternFill(fill_type=None)
                    
                    cancelled_appointments.append({
                        'date': cell_date_str,
                        'time': str(cell_time)
                    })
                    found = True
            
            if not found:
                return False, f"No reservation found for {patient_name} with {doctor_name}"
            
            # Save the workbook
            wb.save(self.excel_path)
            wb.close()
            
            # Create success message
            if len(cancelled_appointments) == 1:
                appt = cancelled_appointments[0]
                message = f"✅ Appointment cancelled successfully!\n\nDoctor: {doctor_name}\nDate: {appt['date']}\nTime: {appt['time']}\nPatient: {patient_name}"
            else:
                message = f"✅ {len(cancelled_appointments)} appointments cancelled for {patient_name} with {doctor_name}"
            
            return True, message
            
        except Exception as e:
            return False, f"Error cancelling appointment: {str(e)}"
    
    def search_appointments(
        self,
        patient_name: Optional[str] = None,
        doctor_name: Optional[str] = None,
        date: Optional[str] = None
    ) -> List[Dict]:
        """
        Search for appointments based on criteria
        
        Args:
            patient_name: Patient's name (optional)
            doctor_name: Doctor's name (optional)
            date: Date to search (optional)
            
        Returns:
            List of matching appointments
        """
        results = []
        
        # Determine which sheets to search
        sheets_to_search = [doctor_name] if doctor_name and doctor_name in self.doctor_sheets else self.doctor_sheets
        
        for sheet_name in sheets_to_search:
            df = pd.read_excel(self.excel_path, sheet_name=sheet_name)
            
            # Filter for reserved appointments
            reserved_df = df[df['Status'] == 'Reserved'].copy()
            
            # Apply filters
            if patient_name:
                reserved_df = reserved_df[reserved_df['Patient_Name'] == patient_name]
            
            if date:
                reserved_df['Date'] = pd.to_datetime(reserved_df['Date'])
                target_date = pd.to_datetime(date)
                reserved_df = reserved_df[reserved_df['Date'] == target_date]
            
            # Add results
            for _, row in reserved_df.iterrows():
                results.append({
                    'doctor': sheet_name,
                    'date': row['Date'].strftime('%Y-%m-%d') if isinstance(row['Date'], pd.Timestamp) else str(row['Date']),
                    'time': row['Time'],
                    'patient_name': row['Patient_Name'],
                    'phone': row['Phone'],
                    'status': row['Status']
                })
        
        return results
    
    def get_patient_info(self, patient_name: str) -> Optional[Dict]:
        """Get patient information from the Patients sheet"""
        try:
            df = pd.read_excel(self.excel_path, sheet_name='Patients')
            patient_row = df[df['Full_Name'] == patient_name]
            
            if patient_row.empty:
                return None
            
            patient = patient_row.iloc[0]
            return {
                'patient_id': patient['Patient_ID'],
                'full_name': patient['Full_Name'],
                'date_of_birth': str(patient['Date_of_Birth']),
                'gender': patient['Gender'],
                'phone': patient['Phone'],
                'address': patient['Address'],
                'doctor': patient['Doctor']
            }
        except Exception as e:
            print(f"Error getting patient info: {e}")
            return None
